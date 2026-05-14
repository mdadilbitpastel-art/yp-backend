"""Single-tab Excel cheat-sheet for the frontend team.

Produces api_frontend_guide.xlsx with one beautifully styled sheet:
  * Project header banner
  * Quick-reference block (base URL, auth, format, conventions)
  * One row per endpoint, page-color-coded, showing method/URL/response shape/notes

Run with:  python build_api_single_sheet.py
"""
from __future__ import annotations

import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "api_frontend_guide.xlsx")

# ---------- palette ------------------------------------------------------
INDIGO  = "4F46E5"
INK     = "1E293B"
SLATE   = "475569"
MUTED   = "94A3B8"
BORDER  = "E2E8F0"
WHITE   = "FFFFFF"

# soft per-page tints
PAGE_TINT = {
    "Home":      "EEF2FF",
    "About Us":  "ECFDF5",
    "Schools":   "EFF6FF",
    "Employers": "FFFBEB",
    "Partners":  "FCE7F3",
    "Events":    "F0F9FF",
    "Insight":   "F5F3FF",
}
PAGE_ACCENT = {
    "Home":      "4F46E5",
    "About Us":  "059669",
    "Schools":   "0284C7",
    "Employers": "B45309",
    "Partners":  "BE185D",
    "Events":    "0369A1",
    "Insight":   "7C3AED",
}

thin = Side(style="thin", color=BORDER)
box = Border(left=thin, right=thin, top=thin, bottom=thin)


# ---------- data ---------------------------------------------------------
# (Page, Section, URL, Top-level response keys, Notes)
ENDPOINTS = [
    # ---- Home ----
    ("Home", "Page payload",     "/api/home/",
        "header, hero, features, about, network, talent_pool, apply, social_media, testimonials, app, footer",
        "Combined payload. Use when first-loading the homepage to avoid 11 round-trips."),
    ("Home", "Header",           "/api/home/header/",
        "logo, button_text, button_url, tabs[]",
        "Tabs sorted server-side. tabs[].url may be a path or absolute URL."),
    ("Home", "Footer",           "/api/home/footer/",
        "logo, title, address, email, copyright_text, links[]",
        "links[].url is always an absolute URL."),
    ("Home", "Hero",             "/api/home/hero/",
        "id, title, description, highlight_text, primary_button{text,url}, secondary_button{text,url}, rating, bottom_note, images[]",
        "rating is a decimal (0–5) or null. images[] are background images from SectionImage."),
    ("Home", "Features",         "/api/home/features/",
        "title, description, button_text, cards[]",
        "button_text is shared across all cards. Each card has position, title, icon, button_url."),
    ("Home", "About",            "/api/home/about/",
        "label, title, description, images[], primary_button{text,url}, secondary_button{text,url}",
        "Mission / About section."),
    ("Home", "Network",          "/api/home/network/",
        "title, video_url, stats[]",
        "stats[] = selected Statistic rows {value, label}. video_url is a CDN-style absolute URL or null."),
    ("Home", "Talent Pool",      "/api/home/talent-pool/",
        "label, title, subtitle, description, images[], primary_button{text,url}, secondary_button{text,url}",
        "Same shape family as Hero/About."),
    ("Home", "Apply",            "/api/home/apply/",
        "title, subtitle, companies[], employer_logos[], bottom_button{text,url}",
        "companies[] = inline company cards (label/title/description/buttons/image/logo). employer_logos[] = pickers from Data Mgmt → Employers."),
    ("Home", "Social Media",     "/api/home/social-media/",
        "label, heading, subtitle, cards[]",
        "cards[] = picked SocialMediaIcon rows {position, name, icon}."),
    ("Home", "Testimonials",     "/api/home/testimonials/",
        "title, images[], users[]",
        "users[] = {position, name, profile_image, message}. images[] are background images."),
    ("Home", "App",              "/api/home/app/",
        "title, description, buttons[], bottom_note, images[]",
        "Up to 3 buttons in buttons[] with {position, text, url}."),

    # ---- About Us ----
    ("About Us", "Page payload", "/api/about-us/",
        "hero, mission, founder, values, journey, pledge, team, community, social_media",
        "Combined payload — every section in one call."),
    ("About Us", "Hero",         "/api/about-us/hero/",
        "id, label, title, description, images[]",
        "Top-of-page hero."),
    ("About Us", "Mission",      "/api/about-us/mission/",
        "label, title, description, images[], stats[]",
        "stats[] = selected Statistic rows {value, label}."),
    ("About Us", "Founder",      "/api/about-us/founder/",
        "label, founder_name, designation, description, founder_message, images[], button{text,url}",
        "founder_message is a personal quote; description is the longer bio."),
    ("About Us", "Values",       "/api/about-us/values/",
        "label, title, subtitle, cards[]",
        "cards[] = {position, icon, label, note}."),
    ("About Us", "Journey",      "/api/about-us/journey/",
        "label, title, subtitle, cards[]",
        "cards[] = {position, image, title, description} — milestone cards."),
    ("About Us", "Pledge",       "/api/about-us/pledge/",
        "label, title, description, images[]",
        "Simple text + images section."),
    ("About Us", "Team",         "/api/about-us/team/",
        "label, title, subtitle, members[]",
        "members[] = {position, profile_image, name, designation, email_url, view_profile_text, view_profile_url}. email_url is a mailto: URL."),
    ("About Us", "Community",    "/api/about-us/community/",
        "label, title, subtitle, cards[]",
        "cards[] = {position, image, name, description, button{text,url}}."),
    ("About Us", "Social Media", "/api/about-us/social-media/",
        "label, heading, subtitle, cards[]",
        "Separate selection from Home — different icon set possible."),

    # ---- Schools ----
    ("Schools", "Page payload",  "/api/schools/",
        "hero, help, employer, benchmark, subscribe, faq",
        "Combined payload."),
    ("Schools", "Hero",          "/api/schools/hero/",
        "label, title, description, primary_button{text,url}, secondary_button{text,url}, images[]",
        "Standard hero shape."),
    ("Schools", "Help",          "/api/schools/help/",
        "label, title, cards[]",
        "cards[] = {position, title, description}."),
    ("Schools", "Employer",      "/api/schools/employer/",
        "label, title, description, button{text,url}, employers[]",
        "employers[] = picked Employer rows {position, name, logo}."),
    ("Schools", "Benchmark",     "/api/schools/benchmark/",
        "label, title, description, cards[]",
        "cards[] = {position, title, description}."),
    ("Schools", "Subscribe",     "/api/schools/subscribe/",
        "label, title, description, button{text,url}, fields[], images[]",
        "fields[] = dynamic form fields {position, field_name, placeholder} the FE renders. No POST endpoint yet — see Notes block."),
    ("Schools", "FAQ",           "/api/schools/faq/",
        "label, title, description, items[]",
        "items[] = {position, question, answer}."),

    # ---- Employers ----
    ("Employers", "Page payload","/api/employers/",
        "hero, network, mission, offer, events",
        "Combined payload. network is shared with Home."),
    ("Employers", "Hero",        "/api/employers/hero/",
        "label, title, description, primary_button{text,url}, secondary_button{text,url}, images[]",
        "Standard hero shape."),
    ("Employers", "Network",     "/api/employers/network/",
        "title, video_url, stats[]",
        "Shared singleton with Home → /api/home/network/."),
    ("Employers", "Mission",     "/api/employers/mission/",
        "label, title, description, button{text,url}, points[], images[]",
        "points[] = bullet list {position, text}."),
    ("Employers", "Offer",       "/api/employers/offer/",
        "label, title, description, cards[]",
        "cards[] = {position, icon, title, description}."),
    ("Employers", "Events",      "/api/employers/events/",
        "label, title, description, button{text,url}, images[]",
        "images[] = {position, image} — teaser image gallery."),

    # ---- Partners ----
    ("Partners", "Page payload", "/api/partners/",
        "hero, partner, family, review, founder",
        "Combined payload."),
    ("Partners", "Hero",         "/api/partners/hero/",
        "label, title, description, stats[]",
        "stats[] = selected Statistic rows {value, label}."),
    ("Partners", "Partner",      "/api/partners/partner/",
        "search_placeholder, explore_button_text, categories[], employers[]",
        "Directory: render search input with placeholder, categories[] as filter chips, employers[] as cards."),
    ("Partners", "Family",       "/api/partners/family/",
        "label, title, description, employers[], load_more_button{text,url}",
        "load_more_button gives both text and target URL."),
    ("Partners", "Review",       "/api/partners/review/",
        "label, title, cards[]",
        "cards[] = {position, name, designation, message}."),
    ("Partners", "Founder",      "/api/partners/founder/",
        "label, title, description, primary_button{text,url}, secondary_button{text,url}, images[]",
        "Hero-style founder block."),

    # ---- Events ----
    ("Events", "Page payload",   "/api/events/",
        "hero, featured, upcoming, missed, submit",
        "Combined payload."),
    ("Events", "Hero",           "/api/events/hero/",
        "label, title, description, primary_button{text,url}, secondary_button{text,url}",
        "No images on this hero — buttons only."),
    ("Events", "Featured",       "/api/events/featured/",
        "label, datetime_label, title, description, category_label, button{text,url}, images[]",
        "datetime_label is free-text (e.g. 'Fri 15 May — 10:00 BST'). Do not parse server-side."),
    ("Events", "Upcoming",       "/api/events/upcoming/",
        "label, title, card_button_text, categories[], cards[]",
        "card_button_text is shared across cards[]. cards[] has image, label, title, description, years_label, price_label, button_url."),
    ("Events", "Missed",         "/api/events/missed/",
        "label, title, description, card_button_text, cards[]",
        "cards[] = {position, video, title, date_label, button_url}. video is an absolute URL or null."),
    ("Events", "Submit",         "/api/events/submit/",
        "label, title, description, button{text,url}, images[]",
        "Submit-an-event teaser. No POST endpoint yet — see Notes block."),

    # ---- Insight ----
    ("Insight", "Page payload",  "/api/insight/",
        "hero, founder, article, lane, subscribe",
        "Combined payload."),
    ("Insight", "Hero",          "/api/insight/hero/",
        "label, title, description, search_placeholder",
        "Render search input with search_placeholder."),
    ("Insight", "Founder",       "/api/insight/founder/",
        "label_1, label_2, date_label, title, description, meta_data, button{text,url}, categories[], images[]",
        "meta_data is byline / read-time line, free-text. categories[] = {position, name}."),
    ("Insight", "Article",       "/api/insight/article/",
        "title, card_button_text, cards[]",
        "card_button_text shared across cards[]. cards[] has label, image, date_label, title, description, tag, button_url."),
    ("Insight", "Lane",          "/api/insight/lane/",
        "label, title, lanes[]",
        "lanes[] = {position, name, article_count, url}."),
    ("Insight", "Subscribe",     "/api/insight/subscribe/",
        "label, title, description, email_placeholder, button{text,url}, bottom_note, images[]",
        "Newsletter section. No POST endpoint yet — see Notes block."),
]


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "API Guide"

    # column widths (8 columns)
    widths = [4, 12, 16, 9, 38, 56, 50, 6]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---------- title banner ----------
    ws.merge_cells("A1:H1")
    title = ws["A1"]
    title.value = "Young Professionals — Frontend API Guide"
    title.font = Font(name="Calibri", size=18, bold=True, color=WHITE)
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    title.fill = PatternFill("solid", fgColor=INDIGO)
    ws.row_dimensions[1].height = 38

    ws.merge_cells("A2:H2")
    subtitle = ws["A2"]
    subtitle.value = ("One-stop reference for hitting the backend. Every section is one row — "
                      "URL, response shape, gotchas. Read top-to-bottom; group by colour for each page.")
    subtitle.font = Font(name="Calibri", size=11, italic=True, color=SLATE)
    subtitle.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    subtitle.fill = PatternFill("solid", fgColor="F8FAFC")
    ws.row_dimensions[2].height = 28

    # ---------- quick-reference info block ----------
    info_rows = [
        ("Base URL",       "https://<host>/api/   (set <host> per environment — dev / staging / prod)"),
        ("Auth",           "None. All endpoints are public."),
        ("Method",         "GET only. There are no POST / PUT / PATCH / DELETE endpoints in the public API right now."),
        ("Response",       "application/json   ·   never paginated   ·   list keys may be []   ·   file URLs may be null"),
        ("Ordering",       "Lists are pre-sorted server-side; each item also carries a 1-based 'position' field."),
        ("Forms in the UI", "Schools Subscribe, Insight Subscribe, Events Submit — backend POST endpoints are NOT built yet. Frontend can render the inputs, but submission targets need to be added later (see last rows for which keys to use)."),
    ]
    row = 3
    for label, val in info_rows:
        ws.cell(row=row, column=1, value=label)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row=row, column=3, value=val)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)

        l = ws.cell(row=row, column=1)
        l.font = Font(name="Calibri", size=11, bold=True, color=INDIGO)
        l.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        l.fill = PatternFill("solid", fgColor="EEF2FF")
        l.border = box

        v = ws.cell(row=row, column=3)
        v.font = Font(name="Calibri", size=11, color=INK)
        v.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        v.fill = PatternFill("solid", fgColor=WHITE)
        v.border = box
        ws.row_dimensions[row].height = 22
        row += 1

    # spacer
    row += 1

    # ---------- legend ----------
    legend_row = row
    ws.cell(row=legend_row, column=1, value="Page colour key")
    ws.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=2)
    ws.cell(row=legend_row, column=1).font = Font(name="Calibri", size=11, bold=True, color=SLATE)
    ws.cell(row=legend_row, column=1).alignment = Alignment(horizontal="left", indent=1, vertical="center")

    col = 3
    for page, tint in PAGE_TINT.items():
        c = ws.cell(row=legend_row, column=col, value=page)
        c.fill = PatternFill("solid", fgColor=tint)
        c.font = Font(name="Calibri", size=10, bold=True, color=PAGE_ACCENT[page])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = box
        col += 1
    ws.row_dimensions[legend_row].height = 22
    row = legend_row + 2

    # ---------- main table header ----------
    headers = ["#", "Page", "Section", "Method", "URL", "Response top-level keys", "Notes / gotchas", "Auth"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=INDIGO)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        cell.border = box
    ws.row_dimensions[row].height = 26
    ws.freeze_panes = ws.cell(row=row + 1, column=1)
    table_header_row = row
    row += 1

    # ---------- data rows ----------
    for i, (page, section, url, keys, notes) in enumerate(ENDPOINTS, start=1):
        tint = PAGE_TINT.get(page, WHITE)
        accent = PAGE_ACCENT.get(page, SLATE)

        values = [i, page, section, "GET", url, keys, notes, "Public"]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = box
            cell.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
            if c == 1:
                cell.font = Font(name="Calibri", size=11, color=MUTED, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill("solid", fgColor=WHITE)
            elif c == 2:  # page badge
                cell.font = Font(name="Calibri", size=11, bold=True, color=accent)
                cell.fill = PatternFill("solid", fgColor=tint)
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            elif c == 3:  # section name
                cell.font = Font(name="Calibri", size=11, bold=True, color=INK)
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                cell.fill = PatternFill("solid", fgColor=WHITE)
            elif c == 4:  # method pill
                cell.value = "GET"
                cell.font = Font(name="Calibri", size=11, bold=True, color="065F46")
                cell.fill = PatternFill("solid", fgColor="D1FAE5")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c == 5:  # URL
                cell.font = Font(name="Consolas", size=11, color=INK)
                cell.alignment = Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True)
                cell.fill = PatternFill("solid", fgColor=WHITE)
            elif c == 6:  # response keys
                cell.font = Font(name="Consolas", size=10, color=INK)
                cell.alignment = Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True)
                cell.fill = PatternFill("solid", fgColor="F8FAFC")
            elif c == 7:  # notes
                cell.font = Font(name="Calibri", size=10, color=SLATE)
                cell.alignment = Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True)
                cell.fill = PatternFill("solid", fgColor=WHITE)
            elif c == 8:  # auth
                cell.font = Font(name="Calibri", size=10, color="065F46", bold=True)
                cell.fill = PatternFill("solid", fgColor="ECFDF5")
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # estimate row height from longest wrapping content
        widest_keys = len(keys) // 60 + keys.count(",") // 4 + 1
        widest_notes = len(notes) // 50 + 1
        ws.row_dimensions[row].height = max(28, 16 * max(widest_keys, widest_notes))
        row += 1

    # ---------- footer note ----------
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    f = ws.cell(row=row, column=1)
    f.value = (
        "TIP — On any page that has multiple sections, hit the 'Page payload' endpoint once (e.g. GET /api/home/) "
        "instead of firing 11 separate requests. The combined response contains every section already keyed by name."
    )
    f.font = Font(name="Calibri", size=10, italic=True, color=SLATE)
    f.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    f.fill = PatternFill("solid", fgColor="FEF3C7")
    f.border = box
    ws.row_dimensions[row].height = 36

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
