"""Generate a polished Excel workbook documenting the project's REST API.

Tabs produced (one .xlsx file at api_specification.xlsx):
  1. README           — purpose, base URL, auth, conventions, status codes
  2. Endpoints        — every endpoint with method, URL, auth, content-type, purpose
  3. Response Schema  — one row per field per endpoint with required/optional
  4. Write Endpoints  — currently none; lists frontend forms that need backend
  5. Sample Responses — one example JSON per endpoint
  6. Field Types      — reference for the type vocabulary used in the schema

Run with:  python build_api_xlsx.py
"""
from __future__ import annotations

import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


OUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "api_specification.xlsx")


# ---------- styling helpers ----------------------------------------------
INDIGO  = "4F46E5"
INDIGO_TINT = "EEF2FF"
SLATE   = "475569"
BORDER  = "E2E8F0"
HEAD_FG = "FFFFFF"

thin = Side(style="thin", color=BORDER)
box  = Border(left=thin, right=thin, top=thin, bottom=thin)

header_font = Font(name="Calibri", size=11, bold=True, color=HEAD_FG)
header_fill = PatternFill("solid", fgColor=INDIGO)
header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
body_font = Font(name="Calibri", size=11)
body_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
zebra_fill = PatternFill("solid", fgColor="F8FAFC")
section_font = Font(name="Calibri", size=12, bold=True, color=INDIGO)


def style_header_row(ws: Worksheet, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = box


def set_col_widths(ws: Worksheet, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def fill_row(ws: Worksheet, row: int, values: list, zebra: bool = False) -> None:
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = body_font
        cell.alignment = body_align
        cell.border = box
        if zebra:
            cell.fill = zebra_fill


# ---------- README sheet --------------------------------------------------
def write_readme(wb: Workbook) -> None:
    ws = wb.create_sheet("README", 0)
    set_col_widths(ws, [22, 90])

    title = ws.cell(row=1, column=1, value="Young Professionals — Public REST API specification")
    title.font = Font(name="Calibri", size=16, bold=True, color=INDIGO)
    ws.row_dimensions[1].height = 26
    ws.merge_cells("A1:B1")

    sub = ws.cell(row=2, column=1, value=(
        "Authoritative reference for the frontend team. Every section is a tab. "
        "Open the 'Endpoints' tab for the full URL list and the 'Response Schema' "
        "tab for field-by-field detail."
    ))
    sub.font = Font(name="Calibri", size=11, italic=True, color=SLATE)
    sub.alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:B2")
    ws.row_dimensions[2].height = 34

    rows = [
        ("Base URL",          "https://<host>/api/  (set <host> per environment — dev, staging, prod)"),
        ("Auth",              "None. All endpoints are public read-only GET."),
        ("Method support",    "Only GET. No POST/PUT/PATCH/DELETE are exposed. Content is edited via the dashboard CMS at /dashboard/."),
        ("Response type",     "application/json; charset=utf-8"),
        ("Pagination",        "Not used — list payloads are always returned in full, already filtered + ordered server-side."),
        ("CORS",              "Configure ALLOWED_HOSTS / CORS allowed origins in Django settings per environment."),
        ("Image / video URLs","Absolute URLs built from request scheme + host. Frontend should treat them as opaque CDN-style URLs."),
        ("Null vs empty",     "Files (image, icon, video, logo): null when not uploaded. Text fields: empty string when blank."),
        ("Ordering",          "All list items include a 1-based 'position' field reflecting dashboard order. Sort by it client-side if needed."),
        ("Error model",       "Standard DRF: 404 if a singleton has no row yet; 400 with {detail: ...} on misuse. No public POST so no validation errors to handle."),
        ("Status codes",      "200 OK on success. 404 if requested singleton missing (rare — most are auto-created on first dashboard save)."),
        ("Caching headers",   "None set explicitly. Add a CDN/edge cache layer if needed."),
        ("Singletons",        "Each section (Hero, Footer, etc.) is a singleton model loaded via .load() — one row per section globally."),
    ]
    row = 4
    for key, val in rows:
        k = ws.cell(row=row, column=1, value=key)
        v = ws.cell(row=row, column=2, value=val)
        k.font = Font(name="Calibri", size=11, bold=True, color=SLATE)
        v.font = body_font
        v.alignment = body_align
        k.alignment = Alignment(vertical="top")
        k.border = box
        v.border = box
        ws.row_dimensions[row].height = 36
        row += 1

    row += 1
    note = ws.cell(row=row, column=1, value="How to read 'Response Schema'")
    note.font = section_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1
    fill_row(ws, row, [
        "Required",
        "Always present in the response. For lists, the key is always there but the array may be empty []."
    ])
    row += 1
    fill_row(ws, row, [
        "Nullable",
        "May be null (typically for file/image/video fields when no upload exists)."
    ], zebra=True)
    row += 1
    fill_row(ws, row, [
        "Blank-able",
        "May be an empty string '' (text fields with allow_blank in the serializer)."
    ])
    row += 1


# ---------- endpoint catalogue ---------------------------------------------
ENDPOINTS = [
    # Home
    ("Home", "Page payload",     "/api/home/",                "Combined Home payload (header + every section + footer)."),
    ("Home", "Header",           "/api/home/header/",         "Site header (logo, CTA button, tabs)."),
    ("Home", "Footer",           "/api/home/footer/",         "Site footer (logo, address, email, copyright, links)."),
    ("Home", "Hero",             "/api/home/hero/",           "Hero section (headline + 2 buttons + rating + images)."),
    ("Home", "Features",         "/api/home/features/",       "Feature cards section."),
    ("Home", "About",            "/api/home/about/",          "About/Mission section."),
    ("Home", "Network",          "/api/home/network/",        "Network section (title + video + selected statistics)."),
    ("Home", "Talent Pool",      "/api/home/talent-pool/",    "Talent Pool section."),
    ("Home", "Apply",            "/api/home/apply/",          "Apply section (cards + selected employer logos)."),
    ("Home", "Social Media",     "/api/home/social-media/",   "Social media icons section."),
    ("Home", "Testimonials",     "/api/home/testimonials/",   "Testimonials section."),
    ("Home", "App",              "/api/home/app/",            "App promo section."),
    # About Us
    ("About Us", "Page payload",  "/api/about-us/",            "Combined About Us payload."),
    ("About Us", "Hero",          "/api/about-us/hero/",       "About Us hero."),
    ("About Us", "Mission",       "/api/about-us/mission/",    "Mission section + stats."),
    ("About Us", "Founder",       "/api/about-us/founder/",    "Founder section."),
    ("About Us", "Values",        "/api/about-us/values/",     "Values cards."),
    ("About Us", "Journey",       "/api/about-us/journey/",    "Journey milestone cards."),
    ("About Us", "Pledge",        "/api/about-us/pledge/",     "Pledge section."),
    ("About Us", "Team",          "/api/about-us/team/",       "Team members."),
    ("About Us", "Community",     "/api/about-us/community/",  "Community cards."),
    ("About Us", "Social Media",  "/api/about-us/social-media/", "About Us social icons."),
    # Schools
    ("Schools", "Page payload",   "/api/schools/",             "Combined Schools payload."),
    ("Schools", "Hero",           "/api/schools/hero/",        "Schools hero."),
    ("Schools", "Help",           "/api/schools/help/",        "Help cards."),
    ("Schools", "Employer",       "/api/schools/employer/",    "Employer logos section."),
    ("Schools", "Benchmark",      "/api/schools/benchmark/",   "Benchmark cards."),
    ("Schools", "Subscribe",      "/api/schools/subscribe/",   "Subscribe section + dynamic form fields config."),
    ("Schools", "FAQ",            "/api/schools/faq/",         "FAQ Q&A items."),
    # Employers
    ("Employers", "Page payload", "/api/employers/",           "Combined Employers payload."),
    ("Employers", "Hero",         "/api/employers/hero/",      "Employers hero."),
    ("Employers", "Network",      "/api/employers/network/",   "Network section (shared with Home)."),
    ("Employers", "Mission",      "/api/employers/mission/",   "Mission with bullet points."),
    ("Employers", "Offer",        "/api/employers/offer/",     "Offer cards."),
    ("Employers", "Events",       "/api/employers/events/",    "Events teaser images."),
    # Partners
    ("Partners", "Page payload",  "/api/partners/",            "Combined Partners payload."),
    ("Partners", "Hero",          "/api/partners/hero/",       "Partners hero with stats."),
    ("Partners", "Partner",       "/api/partners/partner/",    "Partner directory (categories + employer cards)."),
    ("Partners", "Family",        "/api/partners/family/",     "Family section."),
    ("Partners", "Review",        "/api/partners/review/",     "Review cards."),
    ("Partners", "Founder",       "/api/partners/founder/",    "Founder section."),
    # Events
    ("Events", "Page payload",    "/api/events/",              "Combined Events payload."),
    ("Events", "Hero",            "/api/events/hero/",         "Events hero."),
    ("Events", "Featured",        "/api/events/featured/",     "Featured event card."),
    ("Events", "Upcoming",        "/api/events/upcoming/",     "Upcoming events with categories + cards."),
    ("Events", "Missed",          "/api/events/missed/",       "Missed events (video recaps)."),
    ("Events", "Submit",          "/api/events/submit/",       "Submit-an-event teaser section."),
    # Insight
    ("Insight", "Page payload",   "/api/insight/",             "Combined Insight payload."),
    ("Insight", "Hero",           "/api/insight/hero/",        "Insight hero with search placeholder."),
    ("Insight", "Founder",        "/api/insight/founder/",     "Founder spotlight."),
    ("Insight", "Article",        "/api/insight/article/",     "Article cards."),
    ("Insight", "Lane",           "/api/insight/lane/",        "Lanes (pick-your-lane)."),
    ("Insight", "Subscribe",      "/api/insight/subscribe/",   "Subscribe section."),
]


def write_endpoints(wb: Workbook) -> None:
    ws = wb.create_sheet("Endpoints")
    set_col_widths(ws, [4, 12, 18, 10, 38, 10, 16, 60])
    headers = ["#", "Page", "Section", "Method", "URL", "Auth", "Content-Type", "Purpose"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for i, (page, section, url, purpose) in enumerate(ENDPOINTS, start=1):
        row = i + 1
        fill_row(ws, row, [
            i, page, section, "GET", url, "Public", "application/json", purpose
        ], zebra=(i % 2 == 0))


# ---------- response schema --------------------------------------------------
# Each entry: (Field path, Type, Required, Nullable, Blank-able, Description)
# Required = always present in the response object
# Nullable = the value may be null (for files when absent)
# Blank-able = string fields that may be empty ""
F = lambda path, type_, req=True, null=False, blank=False, desc="": (path, type_, "Yes" if req else "No", "Yes" if null else "No", "Yes" if blank else "No", desc)


HEADER_FIELDS = [
    F("logo",         "string (URL)", null=True,  desc="Header logo image."),
    F("button_text",  "string",       blank=True, desc="CTA button label."),
    F("button_url",   "string",       blank=True, desc="CTA URL (internal path or absolute URL)."),
    F("tabs",         "object[]",                 desc="Navigation tabs array (may be empty)."),
    F("tabs[].label", "string",                   desc="Tab display label."),
    F("tabs[].url",   "string",       blank=True, desc="Tab destination URL."),
]

FOOTER_FIELDS = [
    F("logo",            "string (URL)", null=True,  desc="Footer logo."),
    F("title",           "string",       blank=True, desc="Footer title."),
    F("address",         "string",       blank=True, desc="Postal address."),
    F("email",           "string",       blank=True, desc="Contact email."),
    F("copyright_text",  "string",       blank=True, desc="Copyright line."),
    F("links",           "object[]",                 desc="Footer links array (may be empty)."),
    F("links[].label",   "string",                   desc="Link label."),
    F("links[].url",     "string (URL)", blank=True, desc="Link URL."),
]

HOME_HERO = [
    F("id",                    "integer",                  desc="Hero record id."),
    F("title",                 "string",       blank=True, desc="Headline."),
    F("description",           "string",       blank=True, desc="Sub-headline / paragraph."),
    F("highlight_text",        "string",       blank=True, desc="Text portion to highlight."),
    F("primary_button",        "object",                   desc="{text, url}"),
    F("primary_button.text",   "string",       blank=True, desc="Primary CTA text."),
    F("primary_button.url",    "string",       blank=True, desc="Primary CTA URL."),
    F("secondary_button",      "object",                   desc="{text, url}"),
    F("secondary_button.text", "string",       blank=True, desc="Secondary CTA text."),
    F("secondary_button.url",  "string",       blank=True, desc="Secondary CTA URL."),
    F("rating",                "number",       null=True,  desc="Out of 5, decimal. Null if not set."),
    F("bottom_note",           "string",       blank=True, desc="Small note under the buttons."),
    F("images",                "string[]",                 desc="Absolute image URLs from SectionImage."),
]

FEATURES = [
    F("title",                "string", blank=True, desc="Section title."),
    F("description",          "string", blank=True, desc="Section description."),
    F("button_text",          "string", blank=True, desc="Shared CTA shown on every card."),
    F("cards",                "object[]",          desc="Feature cards (may be empty)."),
    F("cards[].position",     "integer",           desc="1-based position."),
    F("cards[].title",        "string", blank=True, desc="Card title."),
    F("cards[].icon",         "string (URL)", null=True, desc="Card icon URL."),
    F("cards[].button_url",   "string (URL)", blank=True, desc="Card-level URL."),
]

ABOUT_HOME = [
    F("label",                  "string", blank=True, desc="Small label / eyebrow."),
    F("title",                  "string", blank=True, desc="Section title."),
    F("description",            "string", blank=True, desc="Section body."),
    F("images",                 "string[]",           desc="Section images."),
    F("primary_button",         "object",             desc="{text, url}"),
    F("primary_button.text",    "string", blank=True, desc="Primary CTA text."),
    F("primary_button.url",     "string", blank=True, desc="Primary CTA URL."),
    F("secondary_button",       "object",             desc="{text, url}"),
    F("secondary_button.text",  "string", blank=True, desc="Secondary CTA text."),
    F("secondary_button.url",   "string", blank=True, desc="Secondary CTA URL."),
]

NETWORK = [
    F("title",     "string",       blank=True, desc="Section title."),
    F("video_url", "string (URL)", null=True,  desc="Background video URL."),
    F("stats",     "object[]",                 desc="Selected Statistic rows ({value, label})."),
    F("stats[].value", "string",               desc="Statistic value, e.g. '1,200+'."),
    F("stats[].label", "string",               desc="Statistic label, e.g. 'Students placed'."),
]

TALENT = [
    F("label",                 "string", blank=True, desc="Eyebrow."),
    F("title",                 "string", blank=True, desc="Title."),
    F("subtitle",              "string", blank=True, desc="Subtitle."),
    F("description",           "string", blank=True, desc="Body copy."),
    F("images",                "string[]",           desc="Section images."),
    F("primary_button.text",   "string", blank=True, desc="Primary CTA text."),
    F("primary_button.url",    "string", blank=True, desc="Primary CTA URL."),
    F("secondary_button.text", "string", blank=True, desc="Secondary CTA text."),
    F("secondary_button.url",  "string", blank=True, desc="Secondary CTA URL."),
]

APPLY = [
    F("title",                            "string", blank=True, desc="Section title."),
    F("subtitle",                         "string", blank=True, desc="Subtitle."),
    F("companies",                        "object[]",           desc="Per-company cards."),
    F("companies[].position",             "integer",            desc="1-based position."),
    F("companies[].label",                "string", blank=True, desc="Tag / eyebrow."),
    F("companies[].title",                "string", blank=True, desc="Company name."),
    F("companies[].description",          "string", blank=True, desc="Description."),
    F("companies[].button_text",          "string", blank=True, desc="CTA text."),
    F("companies[].button_url",           "string (URL)", blank=True, desc="CTA URL."),
    F("companies[].image",                "string (URL)", null=True, desc="Big company image URL."),
    F("companies[].logo",                 "string (URL)", null=True, desc="Small company logo URL."),
    F("employer_logos",                   "object[]",           desc="Picker from Data Management → Employers."),
    F("employer_logos[].position",        "integer",            desc="1-based position."),
    F("employer_logos[].name",            "string", blank=True, desc="Employer name."),
    F("employer_logos[].logo",            "string (URL)", null=True, desc="Logo URL."),
    F("employer_logos[].description",     "string", blank=True, desc="Description."),
    F("employer_logos[].url",             "string (URL)", blank=True, desc="Employer site URL."),
    F("bottom_button.text",               "string", blank=True, desc="Bottom CTA text."),
    F("bottom_button.url",                "string", blank=True, desc="Bottom CTA URL."),
]

SOCIAL_HOME = [
    F("label",            "string", blank=True, desc="Eyebrow."),
    F("heading",          "string", blank=True, desc="Section heading."),
    F("subtitle",         "string", blank=True, desc="Subtitle."),
    F("cards",            "object[]",           desc="Social cards array."),
    F("cards[].position", "integer",            desc="1-based position."),
    F("cards[].name",     "string", blank=True, desc="Platform name."),
    F("cards[].icon",     "string (URL)", null=True, desc="Icon URL."),
]

TESTIMONIALS = [
    F("title",                "string", blank=True, desc="Section title."),
    F("images",               "string[]",           desc="Background images."),
    F("users",                "object[]",           desc="Testimonial users."),
    F("users[].position",     "integer",            desc="1-based position."),
    F("users[].name",         "string", blank=True, desc="Person name."),
    F("users[].profile_image","string (URL)", null=True, desc="Profile photo URL."),
    F("users[].message",      "string", blank=True, desc="Testimonial message."),
]

APP_SECTION = [
    F("title",             "string", blank=True, desc="Section title."),
    F("description",       "string", blank=True, desc="Description body."),
    F("buttons",           "object[]",           desc="Up to 3 buttons."),
    F("buttons[].position","integer",            desc="1-based position."),
    F("buttons[].text",    "string", blank=True, desc="Button text."),
    F("buttons[].url",     "string (URL)", blank=True, desc="Button URL."),
    F("bottom_note",       "string", blank=True, desc="Small note under the buttons."),
    F("images",            "string[]",           desc="Section images (phone, barcode, etc.)."),
]

# About Us
ABOUTUS_HERO = [
    F("id",          "integer",           desc="Hero record id."),
    F("label",       "string", blank=True, desc="Eyebrow."),
    F("title",       "string", blank=True, desc="Hero title."),
    F("description", "string", blank=True, desc="Hero description."),
    F("images",      "string[]",          desc="Hero images."),
]
ABOUTUS_MISSION = [
    F("label",       "string", blank=True),
    F("title",       "string", blank=True),
    F("description", "string", blank=True),
    F("images",      "string[]"),
    F("stats",       "object[]", desc="Selected Statistic rows ({value, label})."),
    F("stats[].value", "string"),
    F("stats[].label", "string"),
]
ABOUTUS_FOUNDER = [
    F("label",            "string", blank=True),
    F("founder_name",     "string", blank=True),
    F("designation",      "string", blank=True),
    F("description",      "string", blank=True),
    F("founder_message",  "string", blank=True),
    F("images",           "string[]"),
    F("button",           "object",  desc="{text, url}"),
    F("button.text",      "string", blank=True),
    F("button.url",       "string", blank=True),
]
ABOUTUS_VALUES = [
    F("label",             "string", blank=True),
    F("title",             "string", blank=True),
    F("subtitle",          "string", blank=True),
    F("cards",             "object[]"),
    F("cards[].position",  "integer"),
    F("cards[].icon",      "string (URL)", null=True),
    F("cards[].label",     "string", blank=True),
    F("cards[].note",      "string", blank=True),
]
ABOUTUS_JOURNEY = [
    F("label",             "string", blank=True),
    F("title",             "string", blank=True),
    F("subtitle",          "string", blank=True),
    F("cards",             "object[]"),
    F("cards[].position",  "integer"),
    F("cards[].image",     "string (URL)", null=True),
    F("cards[].title",     "string", blank=True),
    F("cards[].description","string", blank=True),
]
ABOUTUS_PLEDGE = [
    F("label",       "string", blank=True),
    F("title",       "string", blank=True),
    F("description", "string", blank=True),
    F("images",      "string[]"),
]
ABOUTUS_TEAM = [
    F("label",                       "string", blank=True),
    F("title",                       "string", blank=True),
    F("subtitle",                    "string", blank=True),
    F("members",                     "object[]"),
    F("members[].position",          "integer"),
    F("members[].profile_image",     "string (URL)", null=True),
    F("members[].name",              "string", blank=True),
    F("members[].designation",       "string", blank=True),
    F("members[].email_url",         "string", blank=True, desc="mailto: link."),
    F("members[].view_profile_text", "string", blank=True),
    F("members[].view_profile_url",  "string (URL)", blank=True),
]
ABOUTUS_COMMUNITY = [
    F("label",                  "string", blank=True),
    F("title",                  "string", blank=True),
    F("subtitle",               "string", blank=True),
    F("cards",                  "object[]"),
    F("cards[].position",       "integer"),
    F("cards[].image",          "string (URL)", null=True),
    F("cards[].name",           "string", blank=True),
    F("cards[].description",    "string", blank=True),
    F("cards[].button.text",    "string", blank=True),
    F("cards[].button.url",     "string", blank=True),
]
ABOUTUS_SOCIAL = [
    F("label",             "string", blank=True),
    F("heading",           "string", blank=True),
    F("subtitle",          "string", blank=True),
    F("cards",             "object[]"),
    F("cards[].position",  "integer"),
    F("cards[].name",      "string", blank=True),
    F("cards[].icon",      "string (URL)", null=True),
]

# Schools
SCHOOLS_HERO = [
    F("label",                 "string", blank=True),
    F("title",                 "string", blank=True),
    F("description",           "string", blank=True),
    F("primary_button.text",   "string", blank=True),
    F("primary_button.url",    "string", blank=True),
    F("secondary_button.text", "string", blank=True),
    F("secondary_button.url",  "string", blank=True),
    F("images",                "string[]"),
]
SCHOOLS_HELP = [
    F("label",                 "string", blank=True),
    F("title",                 "string", blank=True),
    F("cards",                 "object[]"),
    F("cards[].position",      "integer"),
    F("cards[].title",         "string", blank=True),
    F("cards[].description",   "string", blank=True),
]
SCHOOLS_EMPLOYER = [
    F("label",                  "string", blank=True),
    F("title",                  "string", blank=True),
    F("description",            "string", blank=True),
    F("button.text",            "string", blank=True),
    F("button.url",             "string", blank=True),
    F("employers",              "object[]"),
    F("employers[].position",   "integer"),
    F("employers[].name",       "string", blank=True),
    F("employers[].logo",       "string (URL)", null=True),
]
SCHOOLS_BENCHMARK = [
    F("label",                 "string", blank=True),
    F("title",                 "string", blank=True),
    F("description",           "string", blank=True),
    F("cards",                 "object[]"),
    F("cards[].position",      "integer"),
    F("cards[].title",         "string", blank=True),
    F("cards[].description",   "string", blank=True),
]
SCHOOLS_SUBSCRIBE = [
    F("label",                  "string", blank=True),
    F("title",                  "string", blank=True),
    F("description",            "string", blank=True),
    F("button.text",            "string", blank=True, desc="Subscribe button text."),
    F("button.url",             "string", blank=True, desc="Subscribe button URL (use this for submissions when backend is built)."),
    F("fields",                 "object[]", desc="Dynamic form fields the FE should render."),
    F("fields[].position",      "integer"),
    F("fields[].field_name",    "string", blank=True),
    F("fields[].placeholder",   "string", blank=True),
    F("images",                 "string[]"),
]
SCHOOLS_FAQ = [
    F("label",            "string", blank=True),
    F("title",            "string", blank=True),
    F("description",      "string", blank=True),
    F("items",            "object[]"),
    F("items[].position", "integer"),
    F("items[].question", "string", blank=True),
    F("items[].answer",   "string", blank=True),
]

# Employers
EMP_HERO = SCHOOLS_HERO  # same shape
EMP_MISSION = [
    F("label",                 "string", blank=True),
    F("title",                 "string", blank=True),
    F("description",           "string", blank=True),
    F("button.text",           "string", blank=True),
    F("button.url",            "string", blank=True),
    F("points",                "object[]"),
    F("points[].position",     "integer"),
    F("points[].text",         "string", blank=True),
    F("images",                "string[]"),
]
EMP_OFFER = [
    F("label",                 "string", blank=True),
    F("title",                 "string", blank=True),
    F("description",           "string", blank=True),
    F("cards",                 "object[]"),
    F("cards[].position",      "integer"),
    F("cards[].icon",          "string (URL)", null=True),
    F("cards[].title",         "string", blank=True),
    F("cards[].description",   "string", blank=True),
]
EMP_EVENTS = [
    F("label",                 "string", blank=True),
    F("title",                 "string", blank=True),
    F("description",           "string", blank=True),
    F("button.text",           "string", blank=True),
    F("button.url",            "string", blank=True),
    F("images",                "object[]"),
    F("images[].position",     "integer"),
    F("images[].image",        "string (URL)", null=True),
]

# Partners
PRT_HERO = [
    F("label",       "string", blank=True),
    F("title",       "string", blank=True),
    F("description", "string", blank=True),
    F("stats",       "object[]"),
    F("stats[].value", "string"),
    F("stats[].label", "string"),
]
PRT_PARTNER = [
    F("search_placeholder",     "string", blank=True),
    F("explore_button_text",    "string", blank=True),
    F("categories",             "object[]"),
    F("categories[].position",  "integer"),
    F("categories[].name",      "string", blank=True),
    F("employers",              "object[]"),
    F("employers[].position",   "integer"),
    F("employers[].name",       "string", blank=True),
    F("employers[].logo",       "string (URL)", null=True),
    F("employers[].description","string", blank=True),
    F("employers[].url",        "string (URL)", blank=True),
]
PRT_FAMILY = [
    F("label",                  "string", blank=True),
    F("title",                  "string", blank=True),
    F("description",            "string", blank=True),
    F("employers",              "object[]"),
    F("employers[].position",   "integer"),
    F("employers[].name",       "string", blank=True),
    F("employers[].logo",       "string (URL)", null=True),
    F("employers[].description","string", blank=True),
    F("employers[].url",        "string (URL)", blank=True),
    F("load_more_button.text",  "string", blank=True),
    F("load_more_button.url",   "string", blank=True),
]
PRT_REVIEW = [
    F("label",                 "string", blank=True),
    F("title",                 "string", blank=True),
    F("cards",                 "object[]"),
    F("cards[].position",      "integer"),
    F("cards[].name",          "string", blank=True),
    F("cards[].designation",   "string", blank=True),
    F("cards[].message",       "string", blank=True),
]
PRT_FOUNDER = [
    F("label",                 "string", blank=True),
    F("title",                 "string", blank=True),
    F("description",           "string", blank=True),
    F("primary_button.text",   "string", blank=True),
    F("primary_button.url",    "string", blank=True),
    F("secondary_button.text", "string", blank=True),
    F("secondary_button.url",  "string", blank=True),
    F("images",                "string[]"),
]

# Events
EVT_HERO = [
    F("label",                 "string", blank=True),
    F("title",                 "string", blank=True),
    F("description",           "string", blank=True),
    F("primary_button.text",   "string", blank=True),
    F("primary_button.url",    "string", blank=True),
    F("secondary_button.text", "string", blank=True),
    F("secondary_button.url",  "string", blank=True),
]
EVT_FEATURED = [
    F("label",          "string", blank=True),
    F("datetime_label", "string", blank=True, desc="Free-text date/time line."),
    F("title",          "string", blank=True),
    F("description",    "string", blank=True),
    F("category_label", "string", blank=True),
    F("button.text",    "string", blank=True),
    F("button.url",     "string", blank=True),
    F("images",         "string[]"),
]
EVT_UPCOMING = [
    F("label",                "string", blank=True),
    F("title",                "string", blank=True),
    F("card_button_text",     "string", blank=True, desc="Shared button text used on every card."),
    F("categories",           "object[]"),
    F("categories[].position","integer"),
    F("categories[].name",    "string", blank=True),
    F("cards",                "object[]"),
    F("cards[].position",     "integer"),
    F("cards[].image",        "string (URL)", null=True),
    F("cards[].label",        "string", blank=True),
    F("cards[].title",        "string", blank=True),
    F("cards[].description",  "string", blank=True),
    F("cards[].years_label",  "string", blank=True),
    F("cards[].price_label",  "string", blank=True),
    F("cards[].button_url",   "string (URL)", blank=True),
]
EVT_MISSED = [
    F("label",               "string", blank=True),
    F("title",               "string", blank=True),
    F("description",         "string", blank=True),
    F("card_button_text",    "string", blank=True),
    F("cards",               "object[]"),
    F("cards[].position",    "integer"),
    F("cards[].video",       "string (URL)", null=True, desc="Recap video URL."),
    F("cards[].title",       "string", blank=True),
    F("cards[].date_label",  "string", blank=True),
    F("cards[].button_url",  "string (URL)", blank=True),
]
EVT_SUBMIT = [
    F("label",                 "string", blank=True),
    F("title",                 "string", blank=True),
    F("description",           "string", blank=True),
    F("button.text",           "string", blank=True),
    F("button.url",            "string", blank=True),
    F("images",                "string[]"),
]

# Insight
INS_HERO = [
    F("label",              "string", blank=True),
    F("title",              "string", blank=True),
    F("description",        "string", blank=True),
    F("search_placeholder", "string", blank=True),
]
INS_FOUNDER = [
    F("label_1",                "string", blank=True),
    F("label_2",                "string", blank=True),
    F("date_label",             "string", blank=True),
    F("title",                  "string", blank=True),
    F("description",            "string", blank=True),
    F("meta_data",              "string", blank=True),
    F("button.text",            "string", blank=True),
    F("button.url",             "string", blank=True),
    F("categories",             "object[]"),
    F("categories[].position",  "integer"),
    F("categories[].name",      "string", blank=True),
    F("images",                 "string[]"),
]
INS_ARTICLE = [
    F("title",                  "string", blank=True),
    F("card_button_text",       "string", blank=True),
    F("cards",                  "object[]"),
    F("cards[].position",       "integer"),
    F("cards[].label",          "string", blank=True),
    F("cards[].image",          "string (URL)", null=True),
    F("cards[].date_label",     "string", blank=True),
    F("cards[].title",          "string", blank=True),
    F("cards[].description",    "string", blank=True),
    F("cards[].tag",            "string", blank=True),
    F("cards[].button_url",     "string (URL)", blank=True),
]
INS_LANE = [
    F("label",                  "string", blank=True),
    F("title",                  "string", blank=True),
    F("lanes",                  "object[]"),
    F("lanes[].position",       "integer"),
    F("lanes[].name",           "string", blank=True),
    F("lanes[].article_count",  "integer"),
    F("lanes[].url",            "string (URL)", blank=True),
]
INS_SUBSCRIBE = [
    F("label",             "string", blank=True),
    F("title",             "string", blank=True),
    F("description",       "string", blank=True),
    F("email_placeholder", "string", blank=True),
    F("button.text",       "string", blank=True),
    F("button.url",        "string", blank=True),
    F("bottom_note",       "string", blank=True),
    F("images",            "string[]"),
]


SCHEMA_BY_URL = {
    "/api/home/header/":         HEADER_FIELDS,
    "/api/home/footer/":         FOOTER_FIELDS,
    "/api/home/hero/":           HOME_HERO,
    "/api/home/features/":       FEATURES,
    "/api/home/about/":          ABOUT_HOME,
    "/api/home/network/":        NETWORK,
    "/api/home/talent-pool/":    TALENT,
    "/api/home/apply/":          APPLY,
    "/api/home/social-media/":   SOCIAL_HOME,
    "/api/home/testimonials/":   TESTIMONIALS,
    "/api/home/app/":            APP_SECTION,

    "/api/about-us/hero/":           ABOUTUS_HERO,
    "/api/about-us/mission/":        ABOUTUS_MISSION,
    "/api/about-us/founder/":        ABOUTUS_FOUNDER,
    "/api/about-us/values/":         ABOUTUS_VALUES,
    "/api/about-us/journey/":        ABOUTUS_JOURNEY,
    "/api/about-us/pledge/":         ABOUTUS_PLEDGE,
    "/api/about-us/team/":           ABOUTUS_TEAM,
    "/api/about-us/community/":      ABOUTUS_COMMUNITY,
    "/api/about-us/social-media/":   ABOUTUS_SOCIAL,

    "/api/schools/hero/":      SCHOOLS_HERO,
    "/api/schools/help/":      SCHOOLS_HELP,
    "/api/schools/employer/":  SCHOOLS_EMPLOYER,
    "/api/schools/benchmark/": SCHOOLS_BENCHMARK,
    "/api/schools/subscribe/": SCHOOLS_SUBSCRIBE,
    "/api/schools/faq/":       SCHOOLS_FAQ,

    "/api/employers/hero/":    EMP_HERO,
    "/api/employers/network/": NETWORK,
    "/api/employers/mission/": EMP_MISSION,
    "/api/employers/offer/":   EMP_OFFER,
    "/api/employers/events/":  EMP_EVENTS,

    "/api/partners/hero/":     PRT_HERO,
    "/api/partners/partner/":  PRT_PARTNER,
    "/api/partners/family/":   PRT_FAMILY,
    "/api/partners/review/":   PRT_REVIEW,
    "/api/partners/founder/":  PRT_FOUNDER,

    "/api/events/hero/":      EVT_HERO,
    "/api/events/featured/":  EVT_FEATURED,
    "/api/events/upcoming/":  EVT_UPCOMING,
    "/api/events/missed/":    EVT_MISSED,
    "/api/events/submit/":    EVT_SUBMIT,

    "/api/insight/hero/":      INS_HERO,
    "/api/insight/founder/":   INS_FOUNDER,
    "/api/insight/article/":   INS_ARTICLE,
    "/api/insight/lane/":      INS_LANE,
    "/api/insight/subscribe/": INS_SUBSCRIBE,
}

PAGE_TOP_KEYS = {
    "/api/home/":       ["header", "hero", "features", "about", "network", "talent_pool", "apply",
                         "social_media", "testimonials", "app", "footer"],
    "/api/about-us/":   ["hero", "mission", "founder", "values", "journey", "pledge", "team",
                         "community", "social_media"],
    "/api/schools/":    ["hero", "help", "employer", "benchmark", "subscribe", "faq"],
    "/api/employers/":  ["hero", "network", "mission", "offer", "events"],
    "/api/partners/":   ["hero", "partner", "family", "review", "founder"],
    "/api/events/":     ["hero", "featured", "upcoming", "missed", "submit"],
    "/api/insight/":    ["hero", "founder", "article", "lane", "subscribe"],
}


def write_schema(wb: Workbook) -> None:
    ws = wb.create_sheet("Response Schema")
    set_col_widths(ws, [12, 38, 32, 18, 10, 10, 12, 60])
    headers = ["Page", "Endpoint", "Field path", "Type", "Required", "Nullable", "Blank-able", "Description"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "C2"

    row = 2
    seen_pages = []
    for page, _sec, url, _purpose in ENDPOINTS:
        # Page payload: list top-level keys pointing to the section endpoints
        if url in PAGE_TOP_KEYS:
            for key in PAGE_TOP_KEYS[url]:
                fill_row(ws, row, [
                    page, url, key, "object", "Yes", "No", "No",
                    f"See response schema for /{url.strip('/').split('/', 1)[0]} → {key}.",
                ], zebra=(row % 2 == 0))
                row += 1
        # Section endpoints with detailed fields
        if url in SCHEMA_BY_URL:
            for (path, type_, req, null, blank, desc) in SCHEMA_BY_URL[url]:
                fill_row(ws, row, [page, url, path, type_, req, null, blank, desc], zebra=(row % 2 == 0))
                row += 1


# ---------- Write endpoints (none today) ---------------------------------
WRITE_NEEDED = [
    # (Page, Form, Suggested method, Suggested URL, Suggested payload, Why it's needed)
    ("Schools",  "Subscribe form",       "POST", "/api/schools/subscribe/submissions/ (suggested)",
        "{ <field_name>: <value>, ... } using fields[].field_name from /api/schools/subscribe/",
        "Frontend renders dynamic fields but currently has nowhere to send the submission."),
    ("Insight",  "Newsletter subscribe", "POST", "/api/insight/subscribe/submissions/ (suggested)",
        "{ email: string }",
        "Frontend collects email via email_placeholder but submission endpoint isn't built."),
    ("Events",   "Submit an event",      "POST", "/api/events/submit/submissions/ (suggested)",
        "{ name, email, event_title, event_date, description, ... }",
        "Submit section advertises a CTA; backend endpoint not implemented."),
    ("Schools",  "Employer / Apply",     "POST", "/api/apply/ (suggested)",
        "{ employer_id: int, applicant_name: string, email: string, message: string, cv: file }",
        "Apply section CTAs currently link out; no internal apply submission endpoint."),
    ("All",      "Contact / general",    "POST", "/api/contact/ (suggested)",
        "{ name, email, message }",
        "No general-purpose contact endpoint exists. Add if any contact form is shown."),
]


def write_write_endpoints(wb: Workbook) -> None:
    ws = wb.create_sheet("Write Endpoints")
    set_col_widths(ws, [14, 28, 10, 46, 60, 60])

    title = ws.cell(row=1, column=1, value="Write endpoints (POST / PUT / PATCH / DELETE)")
    title.font = Font(name="Calibri", size=14, bold=True, color=INDIGO)
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 24

    note = ws.cell(row=2, column=1, value=(
        "Status today: the public REST API exposes GET only. All writes happen inside the "
        "dashboard CMS via Django form posts protected by CSRF — not a public REST surface. "
        "The table below lists frontend-visible forms that would need server-side write "
        "endpoints, with a suggested URL + payload shape the backend team can implement."
    ))
    note.font = Font(name="Calibri", size=11, italic=True, color=SLATE)
    note.alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 50

    headers = ["Page", "Form", "Method", "Suggested URL", "Suggested request payload", "Why it's needed"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h)
    style_header_row(ws, 3, len(headers))
    ws.row_dimensions[3].height = 22
    ws.freeze_panes = "A4"

    row = 4
    for entry in WRITE_NEEDED:
        fill_row(ws, row, list(entry), zebra=(row % 2 == 0))
        row += 1

    row += 1
    sub = ws.cell(row=row, column=1, value="Conventions to follow when these are added")
    sub.font = section_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    rules = [
        "Request body: application/json (or multipart/form-data if file uploads are involved).",
        "Success response: 201 Created with { id, ...created_resource } or 200 OK with { detail: 'ok' }.",
        "Validation errors: 400 with DRF default shape, e.g. { email: ['Enter a valid email address.'] }.",
        "Mandatory fields should be enforced server-side; frontend mirrors them with required attrs / Zod schemas.",
        "Throttle/rate-limit anonymous endpoints (django-rest-framework throttles or upstream WAF).",
        "Send confirmation email asynchronously (Celery / background task) — do not block the request.",
    ]
    for r in rules:
        fill_row(ws, row, [r] + [""] * 5)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1


# ---------- Sample responses -------------------------------------------------
SAMPLES = {
    "/api/home/header/": {
        "logo": "https://api.example.com/media/header/logo.svg",
        "button_text": "Get Started",
        "button_url": "/signup",
        "tabs": [
            {"label": "About", "url": "/about-us/"},
            {"label": "Schools", "url": "/schools/"},
        ],
    },
    "/api/home/footer/": {
        "logo": "https://api.example.com/media/footer/logo.svg",
        "title": "Stay connected",
        "address": "London, UK",
        "email": "hello@example.com",
        "copyright_text": "© 2026 Young Professionals",
        "links": [{"label": "Privacy", "url": "https://example.com/privacy"}],
    },
    "/api/home/hero/": {
        "id": 1,
        "title": "Step into the future",
        "description": "We pair students with great employers.",
        "highlight_text": "future",
        "primary_button": {"text": "Get Started", "url": "/signup"},
        "secondary_button": {"text": "Learn More", "url": "/about-us"},
        "rating": 4.8,
        "bottom_note": "Loved by 12,000+ students",
        "images": ["https://api.example.com/media/hero/bg.jpg"],
    },
    "/api/home/network/": {
        "title": "Our growing network",
        "video_url": "https://api.example.com/media/home/network.mp4",
        "stats": [
            {"value": "1,200+", "label": "Students placed"},
            {"value": "300+",   "label": "Partner schools"},
        ],
    },
    "/api/about-us/team/": {
        "label": "Our Team",
        "title": "Meet the people behind Young Professionals",
        "subtitle": "Mentors, builders, advocates.",
        "members": [
            {
                "position": 1,
                "profile_image": "https://api.example.com/media/data/team/jane.jpg",
                "name": "Jane Doe",
                "designation": "Co-founder",
                "email_url": "mailto:jane@example.com",
                "view_profile_text": "View profile",
                "view_profile_url": "https://linkedin.com/in/jane",
            }
        ],
    },
    "/api/schools/faq/": {
        "label": "FAQ",
        "title": "Frequently Asked Questions",
        "description": "Answers to common questions.",
        "items": [
            {"position": 1, "question": "How does it work?", "answer": "It's simple..."},
        ],
    },
    "/api/events/upcoming/": {
        "label": "Upcoming",
        "title": "Don't miss these",
        "card_button_text": "Book Now",
        "categories": [{"position": 1, "name": "Workshops"}],
        "cards": [
            {
                "position": 1,
                "image": "https://api.example.com/media/events/upcoming/1.jpg",
                "label": "Workshop",
                "title": "Career Bootcamp",
                "description": "Two days of hands-on prep.",
                "years_label": "Years 12+",
                "price_label": "Free",
                "button_url": "https://example.com/register/1",
            }
        ],
    },
    "/api/insight/lane/": {
        "label": "Pick your lane",
        "title": "Choose your interest",
        "lanes": [
            {"position": 1, "name": "Career advice", "article_count": 12,
             "url": "https://example.com/insight/career"},
        ],
    },
}


def write_samples(wb: Workbook) -> None:
    import json as _json
    ws = wb.create_sheet("Sample Responses")
    set_col_widths(ws, [40, 90])

    title = ws.cell(row=1, column=1, value="Sample JSON responses (representative subset)")
    title.font = Font(name="Calibri", size=14, bold=True, color=INDIGO)
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 24

    note = ws.cell(row=2, column=1, value=(
        "Pick GET <endpoint>, this is what comes back. The shape applies to every endpoint of "
        "the same kind — list fields will simply contain more items in production. For full "
        "per-field detail, see 'Response Schema'."
    ))
    note.font = Font(name="Calibri", size=11, italic=True, color=SLATE)
    note.alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:B2")
    ws.row_dimensions[2].height = 36

    headers = ["Endpoint", "Sample JSON response"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h)
    style_header_row(ws, 3, len(headers))
    ws.row_dimensions[3].height = 22
    ws.freeze_panes = "A4"

    row = 4
    for url, sample in SAMPLES.items():
        text = _json.dumps(sample, indent=2)
        fill_row(ws, row, [url, text], zebra=(row % 2 == 0))
        ws.row_dimensions[row].height = min(20 + text.count("\n") * 14, 360)
        row += 1


# ---------- Field types reference --------------------------------------------
def write_types(wb: Workbook) -> None:
    ws = wb.create_sheet("Field Types")
    set_col_widths(ws, [22, 90])

    title = ws.cell(row=1, column=1, value="Type vocabulary")
    title.font = Font(name="Calibri", size=14, bold=True, color=INDIGO)
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 24

    headers = ["Type label", "What it means in the JSON response"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=2, column=c, value=h)
    style_header_row(ws, 2, len(headers))
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"

    rows = [
        ("string",          "Plain string. May be empty '' if 'Blank-able' = Yes."),
        ("string (URL)",    "String that is always a URL (path or absolute). May be empty if 'Blank-able' = Yes."),
        ("integer",         "Whole number, no decimals."),
        ("number",          "Decimal number (e.g. rating 4.8)."),
        ("object",          "Nested JSON object — drill into the .child fields."),
        ("object[]",        "Array of objects. May be empty []."),
        ("string[]",        "Array of strings (typically image URLs). May be empty []."),
    ]
    row = 3
    for t, desc in rows:
        fill_row(ws, row, [t, desc], zebra=(row % 2 == 0))
        row += 1


# ---------- Driver -----------------------------------------------------------
def build() -> None:
    wb = Workbook()
    # The default sheet
    default = wb.active
    wb.remove(default)

    write_readme(wb)
    write_endpoints(wb)
    write_schema(wb)
    write_write_endpoints(wb)
    write_samples(wb)
    write_types(wb)

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    build()
